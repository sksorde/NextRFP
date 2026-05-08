# app/utils/extractor.py

"""
PHASE 2 — ENTERPRISE DOCUMENT EXTRACTION ENGINE

Supported Formats:
- DOCX
- PDF
- XLSX
- XLS
- TXT
- CSV
- RTF (basic text fallback)
- future-ready for PPTX / SharePoint sync

Capabilities:
1. Multi-format extraction
2. Structured Excel parsing
3. Safe decoding fallback
4. Enterprise document ingestion
5. Knowledge-base normalization
6. Bid library ingestion support

This powers:
RFP parsing + Knowledge Base ingestion
"""

from io import BytesIO

from docx import Document
import pdfplumber
from openpyxl import load_workbook


# =====================================================
# EXCEL EXTRACTION
# =====================================================

def extract_excel_text(file_bytes):
    """
    Extract structured text from:
    - XLSX
    - XLS

    Includes:
    - sheet names
    - row-level values

    Useful for:
    Shipley scoring sheets
    PQQ matrices
    Compliance trackers
    Pricing tables
    Delivery plans
    """

    stream = BytesIO(file_bytes)

    workbook = load_workbook(
        stream,
        data_only=True
    )

    all_text = []

    for sheet in workbook.worksheets:
        all_text.append(
            f"Sheet: {sheet.title}"
        )

        for row in sheet.iter_rows(
            values_only=True
        ):
            row_values = []

            for cell in row:
                if cell is not None:
                    row_values.append(
                        str(cell)
                    )

            if row_values:
                all_text.append(
                    " | ".join(row_values)
                )

    return "\n".join(all_text)


# =====================================================
# PDF EXTRACTION
# =====================================================

def extract_pdf_text(file_bytes):
    """
    Extract text from PDF.

    Handles:
    - multi-page proposals
    - RFP packs
    - compliance appendices
    """

    stream = BytesIO(file_bytes)

    pages_text = []

    with pdfplumber.open(stream) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()

            if page_text:
                pages_text.append(
                    page_text
                )

    return "\n".join(pages_text)


# =====================================================
# DOCX EXTRACTION
# =====================================================

def extract_docx_text(file_bytes):
    """
    Extract text from Word documents.

    Common for:
    - proposal drafts
    - RFP responses
    - delivery plans
    - executive summaries
    """

    stream = BytesIO(file_bytes)

    document = Document(stream)

    paragraphs = []

    for para in document.paragraphs:
        if para.text.strip():
            paragraphs.append(
                para.text.strip()
            )

    return "\n".join(paragraphs)


# =====================================================
# SAFE TEXT FALLBACK
# =====================================================

def extract_plain_text(file_bytes):
    """
    Safe UTF-8 fallback.

    Supports:
    - TXT
    - CSV
    - RTF fallback
    - unknown formats
    """

    return file_bytes.decode(
        "utf-8",
        errors="ignore"
    )


# =====================================================
# MASTER EXTRACTION ROUTER
# =====================================================

def extract_text(file_bytes, filename):
    """
    Enterprise extraction router.

    Auto-detects file type
    and routes extraction safely.
    """

    filename = filename.lower()

    print("\n================================")
    print("DOCUMENT EXTRACTION STARTED")
    print("================================")
    print(f"File: {filename}")

    try:
        # =========================================
        # DOCX
        # =========================================

        if filename.endswith(".docx"):
            print("Detected: DOCX")

            text = extract_docx_text(
                file_bytes
            )

            print(
                f"Extracted chars: {len(text)}"
            )

            return text

        # =========================================
        # PDF
        # =========================================

        if filename.endswith(".pdf"):
            print("Detected: PDF")

            text = extract_pdf_text(
                file_bytes
            )

            print(
                f"Extracted chars: {len(text)}"
            )

            return text

        # =========================================
        # XLSX / XLS
        # =========================================

        if (
            filename.endswith(".xlsx")
            or filename.endswith(".xls")
        ):
            print("Detected: Excel")

            text = extract_excel_text(
                file_bytes
            )

            print(
                f"Extracted chars: {len(text)}"
            )

            return text

        # =========================================
        # TXT / CSV / RTF fallback
        # =========================================

        print(
            "Detected: Plain text fallback"
        )

        text = extract_plain_text(
            file_bytes
        )

        print(
            f"Extracted chars: {len(text)}"
        )

        return text

    except Exception as e:
        print(
            f"Extraction failed: {str(e)}"
        )

        raise Exception(
            f"Document extraction failed: {str(e)}"
        )